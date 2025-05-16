#!/usr/bin/env python3
"""
Usage:
    python enhanced_xlsx_parser.py input.xlsx --output output.json --format json
    python enhanced_xlsx_parser.py input.xlsx --output output.csv --format csv
"""

import os
import sys
import json
import argparse
import csv
from typing import Dict, List, Any, Tuple, Optional, Union, Set, TypedDict
from collections import defaultdict
from io import StringIO

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.styles import Alignment
import pandas as pd


# Type definitions for better code organization
class CellStyle(TypedDict):
    """Type definition for cell style information."""
    font: Dict[str, Any]
    alignment: Dict[str, Any]
    fill: Dict[str, Any]


class MergeInfo(TypedDict):
    """Type definition for merge information."""
    min_row: int
    max_row: int
    min_col: int
    max_col: int
    master_cell: Tuple[int, int]
    value: Any
    width: int
    height: int
    horizontal_alignment: Optional[str]
    vertical_alignment: Optional[str]
    is_official_merge: bool = False
    is_visual_merge: bool = False
    merge_direction: Optional[str] = None
    is_label_column: bool = False
    cell_style: Optional[Dict[str, Any]] = None


class CellData(TypedDict):
    """Type definition for cell data with metadata."""
    value: Any
    is_merged: bool
    is_visual_merge: bool
    style: CellStyle
    merge_info: Optional[Dict[str, Any]]


class EnhancedXLSXParser:

    def __init__(self, file_path: str):
        """
        Initialize the parser with the Excel file path.

        Args:
            file_path: Path to the Excel file
        """
        self.file_path = file_path
        self.workbook = None
        self.merged_cells_map = {}
        self.merged_cells_info = {}
        self.cell_styles = {}
        self.sheet_dimensions = {}
        self.cell_content_analysis = {}
        self.headers = {}

    def load(self) -> bool:
        """
        Load the Excel file.

        Returns:
            bool: True if the file was loaded successfully, False otherwise
        """
        try:
            self.workbook = openpyxl.load_workbook(
                self.file_path,
                data_only=True,  # Get values instead of formulas
                read_only=False  # Need full access for merged cells and styles
            )
            return True
        except Exception as e:
            print(f"Error loading Excel file: {e}", file=sys.stderr)
            return False

    def get_sheet_names(self) -> List[str]:
        """
        Get the names of all sheets in the workbook.

        Returns:
            List[str]: List of sheet names
        """
        if not self.workbook:
            return []
        return self.workbook.sheetnames

    def _process_sheet_metadata(self, sheet_name: str) -> None:
        """
        Process sheet metadata including merged cells, styles, and dimensions.
        Enhanced to better distinguish between actually merged cells and visually
        similar cells that should remain independent.

        Args:
            sheet_name: Name of the sheet to process
        """
        sheet = self.workbook[sheet_name]

        # Initialize data structures for this sheet
        self.merged_cells_map[sheet_name] = {}
        self.merged_cells_info[sheet_name] = {}
        self.cell_styles[sheet_name] = {}
        self.cell_content_analysis[sheet_name] = {}

        # Store sheet dimensions
        self.sheet_dimensions[sheet_name] = {
            'max_row': sheet.max_row,
            'max_col': sheet.max_column
        }

        # Process officially merged cells
        for merged_range in sheet.merged_cells.ranges:
            # Get the top-left cell (master cell) of the merged range
            master_cell_coord = merged_range.min_row, merged_range.min_col
            master_cell = sheet.cell(row=master_cell_coord[0], column=master_cell_coord[1])

            # Store information about this merged range
            self.merged_cells_info[sheet_name][merged_range.coord] = {
                'min_row': merged_range.min_row,
                'max_row': merged_range.max_row,
                'min_col': merged_range.min_col,
                'max_col': merged_range.max_col,
                'master_cell': master_cell_coord,
                'value': master_cell.value,
                'width': merged_range.max_col - merged_range.min_col + 1,
                'height': merged_range.max_row - merged_range.min_row + 1,
                'horizontal_alignment': master_cell.alignment.horizontal,
                'vertical_alignment': master_cell.alignment.vertical,
                'is_official_merge': True
            }

            # Map each cell in the merged range to the master cell
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    self.merged_cells_map[sheet_name][(row, col)] = master_cell_coord

        # First pass: Analyze the structure of the data to detect headers and content patterns
        self._analyze_content_structure(sheet_name)

        # Second pass: Detect visually merged cells with enhanced checks
        self._detect_visually_merged_cells(sheet_name)

        # Process cell styles
        self._process_cell_styles(sheet_name)

    def _analyze_content_structure(self, sheet_name: str) -> None:
        """
        Analyze the content structure of the sheet to identify headers,
        repeating patterns, and distinguish actual data vs. labels.

        Args:
            sheet_name: Name of the sheet to process
        """
        sheet = self.workbook[sheet_name]
        max_row = sheet.max_row
        max_col = sheet.max_column

        # Store content analysis
        content_analysis = {}

        # Look for header rows by checking for rows with mostly non-empty cells followed by data rows
        header_candidates = []
        for row in range(1, min(20, max_row)):  # Check first 20 rows at most
            non_empty_count = 0
            empty_count = 0

            for col in range(1, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.value is not None and str(cell.value).strip():
                    non_empty_count += 1
                else:
                    empty_count += 1

            if non_empty_count > 0 and non_empty_count >= empty_count * 0.8:  # 80% threshold
                # Check if cells are bold or have distinctive formatting
                is_header_formatting = any(
                    sheet.cell(row=row, column=col).font.bold
                    for col in range(1, max_col + 1)
                    if sheet.cell(row=row, column=col).value is not None
                )

                header_candidates.append((row, non_empty_count, is_header_formatting))

        # Identify the most likely header row - prioritize formatting, then content
        if header_candidates:
            # First check for rows with header formatting
            formatted_headers = [h for h in header_candidates if h[2]]
            if formatted_headers:
                likely_header_row = max(formatted_headers, key=lambda h: h[1])[0]
            else:
                # Otherwise take the row with most content
                likely_header_row = max(header_candidates, key=lambda h: h[1])[0]

            # Store header information
            self.headers[sheet_name] = {
                'row': likely_header_row,
                'values': {
                    col: sheet.cell(row=likely_header_row, column=col).value
                    for col in range(1, max_col + 1)
                    if sheet.cell(row=likely_header_row, column=col).value is not None
                }
            }

            # Analyze columns for data types and pattern consistency
            for col in range(1, max_col + 1):
                col_values = []
                data_types = set()
                for row in range(likely_header_row + 1, min(likely_header_row + 21, max_row + 1)):
                    cell_value = sheet.cell(row=row, column=col).value
                    if cell_value is not None:
                        col_values.append(cell_value)
                        data_types.add(type(cell_value).__name__)

                content_analysis[col] = {
                    'header': sheet.cell(row=likely_header_row, column=col).value,
                    'data_types': list(data_types),
                    'sample_count': len(col_values),
                    'sample_values': col_values[:5]  # First 5 values as a sample
                }

        self.cell_content_analysis[sheet_name] = content_analysis

    def _create_vertical_visual_merge(self, sheet_name: str, col: int, start_row: int, end_row: int,
                                 value: Any, col_info: Dict[str, Any]) -> None:
        """
        Create a vertical visual merge for a range of cells.

        Args:
            sheet_name: Name of the sheet
            col: Column number
            start_row: Starting row number
            end_row: Ending row number
            value: Value of the merged cells
            col_info: Column information from content analysis
        """
        if end_row - start_row <= 0:
            return

        sheet = self.workbook[sheet_name]
        range_coord = f"{get_column_letter(col)}{start_row}:{get_column_letter(col)}{end_row}"
        master_cell_coord = (start_row, col)
        master_cell = sheet.cell(row=start_row, column=col)

        # Determine if this is a label column that should be visually merged
        is_label_column = False
        has_numeric_values = any(t in ('int', 'float', 'decimal')
                            for t in col_info.get('data_types', []))
        data_type_diversity = len(set(col_info.get('data_types', [])))

        # Left-aligned text columns with consistent data types are likely labels
        if master_cell.alignment.horizontal == 'left' and not has_numeric_values and data_type_diversity <= 2:
            is_label_column = True

        # Store information about this visual merge
        self.merged_cells_info[sheet_name][range_coord] = {
            'min_row': start_row,
            'max_row': end_row,
            'min_col': col,
            'max_col': col,
            'master_cell': master_cell_coord,
            'value': value,
            'width': 1,
            'height': end_row - start_row + 1,
            'is_visual_merge': True,
            'merge_direction': 'vertical',
            'is_label_column': is_label_column,
            'horizontal_alignment': master_cell.alignment.horizontal,
            'vertical_alignment': master_cell.alignment.vertical,
            'cell_style': self._get_cell_style_summary(master_cell)
        }

        # Only map label columns as visual merges
        if is_label_column:
            for r in range(start_row, end_row + 1):
                self.merged_cells_map[sheet_name][(r, col)] = master_cell_coord

    def _detect_visually_merged_cells(self, sheet_name: str) -> None:
        """
        Improved detection of cells that appear merged but aren't officially merged.
        Takes into account content analysis and contextual information to avoid
        incorrectly treating independent cells as merged.

        Args:
            sheet_name: Name of the sheet to process
        """
        sheet = self.workbook[sheet_name]
        max_row = sheet.max_row
        max_col = sheet.max_column
        content_analysis = self.cell_content_analysis[sheet_name]

        # Get header row if detected
        header_row = self.headers.get(sheet_name, {}).get('row', 1)

        # Process each column (vertical merges)
        for col in range(1, max_col + 1):
            current_value = None
            start_row = None

            # Check if this column is likely to contain data vs. labels
            col_info = content_analysis.get(col, {})
            is_data_column = len(col_info.get('data_types', [])) > 0

            # Scan through rows to find consecutive cells with the same value
            for row in range(1, max_row + 1):
                # Skip header row for visual merge detection
                if row == header_row:
                    continue

                # Skip if cell is already part of an official merge
                if (row, col) in self.merged_cells_map.get(sheet_name, {}):
                    # If we were tracking a group, consider creating a visual merge
                    if start_row is not None and row - start_row > 1:
                        self._create_vertical_visual_merge(sheet_name, col, start_row, row-1,
                                                         current_value, col_info)

                    # Reset tracking
                    current_value = None
                    start_row = None
                    continue

                cell = sheet.cell(row=row, column=col)
                cell_value = cell.value

                # Skip empty cells
                if cell_value is None:
                    # If we were tracking a group and this is a data column,
                    # we should NOT create a visual merge (likely just empty data cells)
                    if start_row is not None and is_data_column:
                        # Don't merge data gaps
                        current_value = None
                        start_row = None

                    # If we were tracking a group and this is a label column,
                    # we might want to create a visual merge
                    elif start_row is not None and row - start_row > 1:
                        self._create_vertical_visual_merge(sheet_name, col, start_row, row-1,
                                                         current_value, col_info)

                    # Reset tracking
                    current_value = None
                    start_row = None
                    continue

                # Handle non-empty cells
                # If this is a new value or the first non-empty cell
                if current_value != cell_value:
                    # If we were tracking a group, evaluate whether to create a visual merge
                    if start_row is not None and row - start_row > 1:
                        self._create_vertical_visual_merge(sheet_name, col, start_row, row-1,
                                                         current_value, col_info)

                    # Start tracking a new group
                    current_value = cell_value
                    start_row = row

            # Handle the last group if it exists
            if start_row is not None and max_row - start_row > 0:
                self._create_vertical_visual_merge(sheet_name, col, start_row, max_row,
                                                 current_value, col_info)

        # Process each row (horizontal merges)
        for row in range(1, max_row + 1):
            current_value = None
            start_col = None

            # Skip header row for visual merge detection as headers often shouldn't be merged
            if row == header_row:
                continue

            # Scan through columns to find consecutive cells with the same value
            for col in range(1, max_col + 1):
                # Skip if this cell is already part of a vertical merge
                if (row, col) in self.merged_cells_map.get(sheet_name, {}):
                    # If we were tracking a group, consider creating a visual merge for it
                    if start_col is not None and col - start_col > 1:
                        self._create_horizontal_visual_merge(sheet_name, row, start_col, col-1,
                                                           current_value, content_analysis)

                    # Reset tracking
                    current_value = None
                    start_col = None
                    continue

                cell = sheet.cell(row=row, column=col)
                cell_value = cell.value

                # Skip empty cells
                if cell_value is None:
                    # If we were tracking a group, consider creating a visual merge
                    if start_col is not None and col - start_col > 1:
                        self._create_horizontal_visual_merge(sheet_name, row, start_col, col-1,
                                                           current_value, content_analysis)

                    # Reset tracking
                    current_value = None
                    start_col = None
                    continue

                # Handle non-empty cells
                # If this is a new value or the first non-empty cell
                if current_value != cell_value:
                    # If we were tracking a group, evaluate whether to create a visual merge
                    if start_col is not None and col - start_col > 1:
                        self._create_horizontal_visual_merge(sheet_name, row, start_col, col-1,
                                                           current_value, content_analysis)

                    # Start tracking a new group
                    current_value = cell_value
                    start_col = col

            # Handle the last group if it exists
            if start_col is not None and max_col - start_col > 0:
                self._create_horizontal_visual_merge(sheet_name, row, start_col, max_col,
                                                   current_value, content_analysis)

    def _create_horizontal_visual_merge(self, sheet_name: str, row: int, start_col: int, end_col: int,
                                   value: Any, content_analysis: Dict[str, Any]) -> None:
        """
        Create a horizontal visual merge for a range of cells.

        Args:
            sheet_name: Name of the sheet
            row: Row number
            start_col: Starting column number
            end_col: Ending column number
            value: Value of the merged cells
            content_analysis: Content analysis information
        """
        if end_col - start_col <= 0:
            return

        sheet = self.workbook[sheet_name]
        range_coord = f"{get_column_letter(start_col)}{row}:{get_column_letter(end_col)}{row}"
        master_cell_coord = (row, start_col)
        master_cell = sheet.cell(row=row, column=start_col)

        # Use intelligent criteria for horizontal merges
        should_create_merge = True

        # Analyze data patterns for each column in the range
        data_types_in_range = set()
        for c in range(start_col, end_col + 1):
            col_info = content_analysis.get(c, {})
            data_types_in_range.update(col_info.get('data_types', []))

        # Don't merge if there's significant data type diversity across columns
        if len(data_types_in_range) > 2:
            should_create_merge = False

        # Check alignment - data fields often have specific alignments
        if master_cell.alignment.horizontal == 'center':
            # Center-aligned cells in a row are often titles or headers, not data
            should_create_merge = True
        elif any('int' in dt or 'float' in dt or 'decimal' in dt for dt in data_types_in_range):
            # Don't merge numeric data horizontally
            should_create_merge = False

        if should_create_merge:
            # Store information about this visual merge
            self.merged_cells_info[sheet_name][range_coord] = {
                'min_row': row,
                'max_row': row,
                'min_col': start_col,
                'max_col': end_col,
                'master_cell': master_cell_coord,
                'value': value,
                'width': end_col - start_col + 1,
                'height': 1,
                'is_visual_merge': True,
                'merge_direction': 'horizontal',
                'horizontal_alignment': master_cell.alignment.horizontal,
                'vertical_alignment': master_cell.alignment.vertical,
                'cell_style': self._get_cell_style_summary(master_cell)
            }

            # Map each cell in the visual merge to the master cell
            for c in range(start_col, end_col + 1):
                # Only map if not already mapped by a vertical merge
                if (row, c) not in self.merged_cells_map.get(sheet_name, {}):
                    self.merged_cells_map[sheet_name][(row, c)] = master_cell_coord

    def _get_cell_style_summary(self, cell: Cell) -> CellStyle:
        """
        Get a summary of cell's style attributes in a serializable format.

        Args:
            cell: The cell to analyze

        Returns:
            CellStyle: Dictionary containing style information
        """
        # Extract color as a string if it exists
        font_color = getattr(cell.font, 'color', None)
        if font_color and hasattr(font_color, 'rgb'):
            font_color = font_color.rgb
        elif font_color and hasattr(font_color, 'theme'):
            font_color = f"theme:{font_color.theme}"

        # Extract fill color as a string if it exists
        fill = getattr(cell, 'fill', None)
        fill_type = getattr(fill, 'fill_type', None)
        start_color = getattr(fill, 'start_color', None)

        if start_color and hasattr(start_color, 'rgb'):
            start_color = start_color.rgb
        elif start_color and hasattr(start_color, 'theme'):
            start_color = f"theme:{start_color.theme}"

        return {
            'font': {
                'bold': getattr(cell.font, 'bold', False),
                'italic': getattr(cell.font, 'italic', False),
                'color': font_color,
            },
            'alignment': {
                'horizontal': getattr(cell.alignment, 'horizontal', None),
                'vertical': getattr(cell.alignment, 'vertical', None),
            },
            'fill': {
                'type': fill_type,
                'start_color': start_color,
            }
        }

    def _process_cell_styles(self, sheet_name: str) -> None:
        """
        Process and store cell styles for the sheet.

        Args:
            sheet_name: Name of the sheet to process
        """
        sheet = self.workbook[sheet_name]
        max_row = self.sheet_dimensions[sheet_name]['max_row']
        max_col = self.sheet_dimensions[sheet_name]['max_col']

        # Process styles for each cell
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                self.cell_styles[sheet_name][(row, col)] = self._get_cell_style_summary(cell)

    def _get_cell_value_and_metadata(self, sheet, row: int, col: int) -> CellData:
        """
        Get the value and metadata for a cell, handling merged cells properly.
        Enhanced to better distinguish between actually merged cells and visually similar cells.

        Args:
            sheet: The worksheet
            row: Row number (1-based)
            col: Column number (1-based)

        Returns:
            CellData: Cell value and metadata
        """
        sheet_name = sheet.title

        # Initialize result with default values
        result: CellData = {
            'value': None,
            'is_merged': False,
            'is_visual_merge': False,
            'style': {'font': {}, 'alignment': {}, 'fill': {}},
            'merge_info': None
        }

        # Check if this is a merged cell (using our map)
        is_merged = False
        master_cell_coord = None
        is_visual_merge = False
        merge_info = None

        # Check for both official and visual merges
        merged_cells_map = self.merged_cells_map.get(sheet_name, {})
        if (row, col) in merged_cells_map:
            is_merged = True
            master_cell_coord = merged_cells_map[(row, col)]

            # Find the merge range info to check if it's a visual merge
            for range_coord, info in self.merged_cells_info[sheet_name].items():
                if info.get('master_cell') == master_cell_coord:
                    merge_info = info
                    is_visual_merge = info.get('is_visual_merge', False)
                    break

        # For visual merges, apply intelligent rules to determine if they should be treated as merged
        if is_merged and is_visual_merge:
            # Get column info from content analysis
            col_info = self.cell_content_analysis.get(sheet_name, {}).get(col, {})

            # For data columns, prefer individual values over merged values
            data_types = col_info.get('data_types', [])
            has_numeric_values = any(t in ('int', 'float', 'decimal') for t in data_types)
            data_type_diversity = len(set(data_types))

            # If this is a column with numeric data or diverse data types,
            # treat it as an independent cell rather than part of a merge
            if has_numeric_values or data_type_diversity > 2:
                is_merged = False
                master_cell_coord = None
                is_visual_merge = False
                merge_info = None

            # Also check the actual cell - if it has a value different from the master cell,
            # it should probably be treated as independent
            elif is_merged:
                master_row, master_col = master_cell_coord
                master_value = sheet.cell(row=master_row, column=master_col).value
                actual_value = sheet.cell(row=row, column=col).value

                if actual_value is not None and actual_value != master_value:
                    # This cell has its own value, so it shouldn't be treated as merged
                    is_merged = False
                    master_cell_coord = None
                    is_visual_merge = False
                    merge_info = None

        # Get the cell value
        if is_merged:
            master_row, master_col = master_cell_coord
            cell = sheet.cell(row=master_row, column=master_col)
            value = cell.value
            if is_visual_merge:
                actual_cell = sheet.cell(row=row, column=col)
                if actual_cell.value is not None:
                    value = actual_cell.value
        else:
            cell = sheet.cell(row=row, column=col)
            value = cell.value

        # Get cell style information
        cell_styles = self.cell_styles.get(sheet_name, {})
        if (row, col) in cell_styles:
            cell_style = cell_styles[(row, col)]
        else:
            cell_style = self._get_cell_style_summary(cell)

        # Prepare the result
        result['value'] = value
        result['is_merged'] = is_merged
        result['is_visual_merge'] = is_visual_merge
        result['style'] = cell_style

        # Add merge information if available
        if merge_info:
            result['merge_info'] = {
                'direction': merge_info.get('merge_direction'),
                'width': merge_info.get('width'),
                'height': merge_info.get('height'),
                'is_official_merge': merge_info.get('is_official_merge', False)
            }

        return result

    def parse_sheet(self, sheet_name: str, include_metadata: bool = False) -> List[List[Any]]:
        """
        Parse a sheet into a 2D array.
        Enhanced with improved handling of merged vs. non-merged cells.

        Args:
            sheet_name: Name of the sheet to parse
            include_metadata: Whether to include metadata about cells

        Returns:
            List[List[Any]]: 2D array of cell values, or cells with metadata if include_metadata is True
        """
        if not self.workbook:
            return []

        if sheet_name not in self.workbook.sheetnames:
            return []

        sheet = self.workbook[sheet_name]

        # Process sheet metadata if not already processed
        if sheet_name not in self.merged_cells_map:
            self._process_sheet_metadata(sheet_name)

        # Get sheet dimensions
        max_row = self.sheet_dimensions[sheet_name]['max_row']
        max_col = self.sheet_dimensions[sheet_name]['max_col']

        # Parse sheet into a 2D array
        data = []
        for row in range(1, max_row + 1):
            row_data = []
            for col in range(1, max_col + 1):
                cell_data = self._get_cell_value_and_metadata(sheet, row, col)

                if include_metadata:
                    row_data.append(cell_data)
                else:
                    row_data.append(cell_data['value'])

            data.append(row_data)

        return data

    def parse_sheet_to_dict(self, sheet_name: str, headers_row: int = 1, include_metadata: bool = False) -> List[Dict[str, Any]]:
        """
        Parse a sheet into a list of dictionaries.
        Enhanced with improved handling of merged vs. non-merged cells.

        Args:
            sheet_name: Name of the sheet to parse
            headers_row: Row number containing headers (1-based)
            include_metadata: Whether to include metadata about cells

        Returns:
            List[Dict[str, Any]]: List of dictionaries with header keys and cell values
        """
        if not self.workbook:
            return []

        if sheet_name not in self.workbook.sheetnames:
            return []

        sheet = self.workbook[sheet_name]

        # Process sheet metadata if not already processed
        if sheet_name not in self.merged_cells_map:
            self._process_sheet_metadata(sheet_name)

        # Get sheet dimensions
        max_row = self.sheet_dimensions[sheet_name]['max_row']
        max_col = self.sheet_dimensions[sheet_name]['max_col']

        # Get headers
        headers = []
        for col in range(1, max_col + 1):
            header_data = self._get_cell_value_and_metadata(sheet, headers_row, col)
            header_value = header_data['value']

            # If header is None, use the column letter as header
            if header_value is None:
                header_value = get_column_letter(col)

            # Ensure unique headers by appending column letter for duplicates
            header = str(header_value)
            if header in headers:
                header = f"{header}_{get_column_letter(col)}"

            headers.append(header)

        # Parse sheet into a list of dictionaries
        data = []
        for row in range(headers_row + 1, max_row + 1):
            row_data = {}
            for col in range(1, max_col + 1):
                # Skip if column exceeds headers length
                if col > len(headers):
                    continue

                header = headers[col - 1]
                cell_data = self._get_cell_value_and_metadata(sheet, row, col)

                if include_metadata:
                    row_data[header] = cell_data
                else:
                    row_data[header] = cell_data['value']

            data.append(row_data)

        return data

    def get_merged_cells_info(self, sheet_name: str) -> Dict:
        """
        Get information about merged cells in the sheet.

        Args:
            sheet_name: Name of the sheet

        Returns:
            Dict: Information about merged cells
        """
        if not self.workbook:
            return {}

        if sheet_name not in self.workbook.sheetnames:
            return {}

        # Process sheet metadata if not already processed
        if sheet_name not in self.merged_cells_info:
            self._process_sheet_metadata(sheet_name)

        return self.merged_cells_info[sheet_name]

    def get_content_analysis(self, sheet_name: str) -> Dict:
        """
        Get the content analysis for the sheet.

        Args:
            sheet_name: Name of the sheet

        Returns:
            Dict: Content analysis information
        """
        if not self.workbook:
            return {}

        if sheet_name not in self.workbook.sheetnames:
            return {}

        # Process sheet metadata if not already processed
        if sheet_name not in self.cell_content_analysis:
            self._process_sheet_metadata(sheet_name)

        return self.cell_content_analysis[sheet_name]

    def get_headers_info(self, sheet_name: str) -> Dict:
        """
        Get information about detected headers in the sheet.

        Args:
            sheet_name: Name of the sheet

        Returns:
            Dict: Header information
        """
        if not self.workbook:
            return {}

        if sheet_name not in self.workbook.sheetnames:
            return {}

        # Process sheet metadata if not already processed
        if sheet_name not in self.headers:
            self._process_sheet_metadata(sheet_name)

        return self.headers.get(sheet_name, {})

    def _json_serializable(self, obj: Any) -> Any:
        """
        Convert objects to JSON serializable format.

        Args:
            obj: Object to convert

        Returns:
            JSON serializable object
        """
        if isinstance(obj, dict):
            return {k: self._json_serializable(v) for k, v in obj.items()}
        elif isinstance(obj, list):
            return [self._json_serializable(item) for item in obj]
        elif isinstance(obj, (set, tuple)):
            return [self._json_serializable(item) for item in obj]
        elif hasattr(obj, '__dict__'):
            # For objects with __dict__, convert to a dict of their attributes
            return {
                "__class__": obj.__class__.__name__,
                **{k: self._json_serializable(v) for k, v in obj.__dict__.items()
                   if not k.startswith('_')}
            }
        else:
            # For other objects, convert to string
            return str(obj)

    def to_json(self, sheet_name: str, headers_row: int = 1, include_metadata: bool = False) -> str:
        """
        Convert a sheet to JSON.

        Args:
            sheet_name: Name of the sheet
            headers_row: Row number containing headers (1-based)
            include_metadata: Whether to include metadata about cells

        Returns:
            str: JSON string
        """
        if not self.workbook:
            return "{}"

        if sheet_name not in self.workbook.sheetnames:
            return "{}"

        # Process sheet metadata if not already processed
        if sheet_name not in self.merged_cells_map:
            self._process_sheet_metadata(sheet_name)

        # Parse sheet to dict
        data = self.parse_sheet_to_dict(sheet_name, headers_row, include_metadata)

        # Create result dictionary
        result = {
            "sheet_name": sheet_name,
            "data": data
        }

        # Include merged cells info and dimensions if metadata is requested
        if include_metadata:
            result["merged_cells"] = self.merged_cells_info[sheet_name]
            result["dimensions"] = self.sheet_dimensions[sheet_name]
            result["content_analysis"] = self.cell_content_analysis[sheet_name]
            result["headers"] = self.headers.get(sheet_name, {})

        # Convert to JSON with improved serialization
        return json.dumps(self._json_serializable(result), indent=2)

    def to_csv(self, sheet_name: str) -> str:
        """
        Convert a sheet to CSV.

        Args:
            sheet_name: Name of the sheet

        Returns:
            str: CSV string
        """
        if not self.workbook:
            return ""

        if sheet_name not in self.workbook.sheetnames:
            return ""

        # Process sheet metadata if not already processed
        if sheet_name not in self.merged_cells_map:
            self._process_sheet_metadata(sheet_name)

        # Parse sheet
        data = self.parse_sheet(sheet_name)

        # Convert to CSV
        csv_buffer = StringIO()
        csv_writer = csv.writer(csv_buffer)

        for row in data:
            # Convert None to empty string
            csv_writer.writerow(['' if cell is None else cell for cell in row])

        return csv_buffer.getvalue()

    def to_pandas(self, sheet_name: str, headers_row: int = 1) -> pd.DataFrame:
        """
        Convert a sheet to pandas DataFrame.

        Args:
            sheet_name: Name of the sheet
            headers_row: Row number containing headers (1-based)

        Returns:
            pd.DataFrame: DataFrame representation of the sheet
        """
        if not self.workbook:
            return pd.DataFrame()

        if sheet_name not in self.workbook.sheetnames:
            return pd.DataFrame()

        # Process sheet metadata if not already processed
        if sheet_name not in self.merged_cells_map:
            self._process_sheet_metadata(sheet_name)

        # Parse sheet
        data = self.parse_sheet(sheet_name)

        # Extract headers (row at headers_row - 1 in the 0-indexed data)
        if headers_row <= len(data):
            headers = data[headers_row - 1]
            # Convert None to empty string in headers
            headers = [f"Col{i+1}" if h is None else str(h) for i, h in enumerate(headers)]

            # Ensure unique headers
            seen_headers = set()
            unique_headers = []
            for header in headers:
                if header in seen_headers:
                    i = 1
                    new_header = f"{header}_{i}"
                    while new_header in seen_headers:
                        i += 1
                        new_header = f"{header}_{i}"
                    header = new_header
                seen_headers.add(header)
                unique_headers.append(header)

            # Create DataFrame with headers
            df = pd.DataFrame(data[headers_row:], columns=unique_headers)
        else:
            # Create DataFrame without headers
            df = pd.DataFrame(data)

        return df

    def export(self, sheet_name: str, output_path: str, format: str = 'json',
               headers_row: int = 1, include_metadata: bool = False) -> bool:
        """
        Export a sheet to a file.

        Args:
            sheet_name: Name of the sheet
            output_path: Path to save the output
            format: Output format ('json', 'csv', or 'excel')
            headers_row: Row number containing headers (1-based)
            include_metadata: Whether to include metadata about cells (only for JSON)

        Returns:
            bool: True if the export was successful, False otherwise
        """
        try:
            if format.lower() == 'json':
                # Export to JSON
                json_data = self.to_json(sheet_name, headers_row, include_metadata)
                with open(output_path, 'w') as f:
                    f.write(json_data)

            elif format.lower() == 'csv':
                # Export to CSV
                csv_data = self.to_csv(sheet_name)
                with open(output_path, 'w', newline='') as f:
                    f.write(csv_data)

            elif format.lower() == 'excel':
                # Export to Excel
                df = self.to_pandas(sheet_name, headers_row)
                df.to_excel(output_path, index=False)

            else:
                print(f"Unsupported export format: {format}", file=sys.stderr)
                return False

            return True

        except Exception as e:
            print(f"Error exporting to {format}: {e}", file=sys.stderr)
            return False


def main():
    """Main function for CLI."""
    parser = argparse.ArgumentParser(description='Enhanced XLSX Parser')
    parser.add_argument('input', help='Input Excel file')
    parser.add_argument('--output', help='Output file path')
    parser.add_argument('--format', choices=['json', 'csv', 'excel'], default='json',
                        help='Output format (default: json)')
    parser.add_argument('--sheet', help='Sheet name to parse (default: first sheet)')
    parser.add_argument('--headers-row', type=int, default=1,
                        help='Row number containing headers (1-based, default: 1)')
    parser.add_argument('--include-metadata', action='store_true',
                        help='Include metadata about cells (only for JSON)')
    parser.add_argument('--analyze-only', action='store_true',
                        help='Only analyze the structure without exporting data')

    args = parser.parse_args()

    # Initialize parser
    parser = EnhancedXLSXParser(args.input)

    # Load the Excel file
    if not parser.load():
        sys.exit(1)

    # Get sheet name
    sheet_names = parser.get_sheet_names()
    if not sheet_names:
        print("No sheets found in the Excel file.", file=sys.stderr)
        sys.exit(1)

    sheet_name = args.sheet if args.sheet else sheet_names[0]
    if sheet_name not in sheet_names:
        print(f"Sheet '{sheet_name}' not found in the Excel file.", file=sys.stderr)
        print(f"Available sheets: {', '.join(sheet_names)}", file=sys.stderr)
        sys.exit(1)

    # If analyze-only flag is set, just print the analysis
    if args.analyze_only:
        # Process sheet metadata
        parser._process_sheet_metadata(sheet_name)

        # Print content analysis
        content_analysis = parser.get_content_analysis(sheet_name)
        print("\n=== Content Analysis ===")
        for col, info in content_analysis.items():
            print(f"Column {col} (Header: {info.get('header', 'None')}):")
            print(f"  Data types: {', '.join(info.get('data_types', []))}")
            print(f"  Sample values: {info.get('sample_values', [])}")
            print()

        # Print headers info
        headers_info = parser.get_headers_info(sheet_name)
        print("\n=== Headers Information ===")
        print(f"Header row: {headers_info.get('row', 'Not detected')}")
        print("Header values:")
        for col, value in headers_info.get('values', {}).items():
            print(f"  Column {col}: {value}")
        print()

        # Print merged cells info
        merged_cells_info = parser.get_merged_cells_info(sheet_name)
        print("\n=== Merged Cells Information ===")
        official_merges = [r for r, info in merged_cells_info.items() if info.get('is_official_merge', False)]
        visual_merges = [r for r, info in merged_cells_info.items() if info.get('is_visual_merge', False)]

        print(f"Official merges: {len(official_merges)}")
        for i, range_coord in enumerate(official_merges[:5], 1):
            info = merged_cells_info[range_coord]
            print(f"  {i}. {range_coord}: {info.get('value')} ({info.get('width')}x{info.get('height')})")
        if len(official_merges) > 5:
            print(f"  ... and {len(official_merges) - 5} more")

        print(f"\nVisual merges: {len(visual_merges)}")
        for i, range_coord in enumerate(visual_merges[:5], 1):
            info = merged_cells_info[range_coord]
            print(f"  {i}. {range_coord}: {info.get('value')} ({info.get('width')}x{info.get('height')}, {info.get('merge_direction')})")
        if len(visual_merges) > 5:
            print(f"  ... and {len(visual_merges) - 5} more")

        sys.exit(0)

    # Export to the specified format
    if not args.output:
        # If no output path is specified, use the input file name with the appropriate extension
        base_name = os.path.splitext(os.path.basename(args.input))[0]
        ext = {'json': '.json', 'csv': '.csv', 'excel': '.xlsx'}[args.format.lower()]
        output_path = f"{base_name}_output{ext}"
    else:
        output_path = args.output

    # Export
    if parser.export(sheet_name, output_path, args.format, args.headers_row, args.include_metadata):
        print(f"Successfully exported to {output_path}")
    else:
        sys.exit(1)


if __name__ == "__main__":
    main()
